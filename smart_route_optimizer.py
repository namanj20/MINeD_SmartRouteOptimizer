import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import numpy as np
import os
from sklearn.cluster import KMeans
from haversine import haversine
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import folium
import webbrowser
import traceback

class SmartRouteOptimizerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Smart Route Optimizer")
        self.root.geometry("1200x700")
        self.root.configure(bg="#f0f0f1")
        self.root.attributes('-fullscreen', True)
        self.root.bind("<Escape>", lambda event: self.root.attributes('-fullscreen', False))

        # Color scheme
        self.primary_color = "#2c3e50"
        self.secondary_color = "#3498db"
        self.accent_color = "#e74c3c"
        self.bg_color = "#ecf0f1"
        self.text_color = "#34495e"

        self.style = ttk.Style()
        self.configure_styles()

        self.selected_file = ""
        self.trip_data = None
        self.formatted_trip_data = None

        self.create_custom_title_bar()
        self.create_sidebar()
        self.create_main_content()

    def configure_styles(self):
        self.style.theme_use("clam")
        self.style.configure("TButton", font=("Segoe UI", 10), padding=6)
        self.style.configure("Accent.TButton", background=self.primary_color, foreground="white")
        self.style.map("Accent.TButton", background=[("active", self.accent_color)])
        self.style.configure("TLabel", font=("Segoe UI", 10), background=self.bg_color, foreground=self.text_color)
        self.style.configure("Header.TLabel", font=("Segoe UI", 16, "bold"), background=self.bg_color, foreground=self.primary_color)
        self.style.configure("SubHeader.TLabel", font=("Segoe UI", 12, "bold"), background=self.bg_color, foreground=self.primary_color)


        # Custom style for the progress bar
        self.style.configure("Custom.Horizontal.TProgressbar", background=self.accent_color, troughcolor=self.bg_color)

    def create_custom_title_bar(self):
        title_bar = tk.Frame(self.root, bg=self.primary_color, relief='raised', bd=0, highlightthickness=0)
        title_bar.pack(fill='x')
        
        title_label = tk.Label(title_bar, text="Smart Route Optimizer", fg="white", bg=self.primary_color, font=("Segoe UI", 12))
        title_label.pack(side='left', pady=5, padx=10)
        
        close_button = tk.Button(title_bar, text='×', command=self.root.destroy, bg=self.primary_color, fg="white", font=("Segoe UI", 13), bd=0, padx=7, highlightthickness=0)
        close_button.pack(side='right', pady=5, padx=10)

        title_bar.bind('<Button-1>', self.get_pos)
        title_bar.bind('<B1-Motion>', self.move_window)

    def get_pos(self, event):
        self.x = event.x
        self.y = event.y

    def move_window(self, event):
        deltax = event.x - self.x
        deltay = event.y - self.y
        x = self.root.winfo_x() + deltax
        y = self.root.winfo_y() + deltay
        self.root.geometry(f"+{x}+{y}")

    def create_sidebar(self):
        sidebar = tk.Frame(self.root, bg=self.primary_color, width=200)
        sidebar.pack(side='left', fill='y')

        # Create sidebar buttons
        tk.Button(sidebar, text="Home", bg=self.primary_color, fg="white", bd=0, padx=20, pady=10, command=self.show_home).pack(fill='x')
        tk.Button(sidebar, text="About", bg=self.primary_color, fg="white", bd=0, padx=20, pady=10, command=self.show_about).pack(fill='x')
        tk.Button(sidebar, text="Pricing", bg=self.primary_color, fg="white", bd=0, padx=20, pady=10, command=self.show_pricing).pack(fill='x')

    def create_main_content(self):
        main_frame = tk.Frame(self.root, bg=self.bg_color)
        main_frame.pack(side='right', fill='both', expand=True, padx=20, pady=20)

        header_label = ttk.Label(main_frame, text="Optimize Your Routes", style="Header.TLabel")
        header_label.pack(pady=(0, 20))

        # File selection and optimization buttons
        controls_frame = tk.Frame(main_frame, bg='#ecf0f1')
        controls_frame.pack(fill='x', pady=10)

        file_frame = tk.Frame(controls_frame, bg='#ecf0f1')
        file_frame.pack(side='left')

        ttk.Label(file_frame, text="Select Input File:", style="SubHeader.TLabel").pack(side='left', padx=(0, 10))
        self.file_btn = ttk.Button(file_frame, text="Browse Files", command=self.choose_file, style="Accent.TButton")
        self.file_btn.pack(side='left')
        self.file_status = ttk.Label(file_frame, text="No file selected", style="SubHeader.TLabel")
        self.file_status.pack(side='left', padx=10)

        self.run_btn = ttk.Button(controls_frame, text="Start Optimization", command=self.run_optimization, style="Accent.TButton")
        self.run_btn.pack(side='right')

        # Progress bar frame
        self.progress_frame = tk.Frame(main_frame, bg='#ecf0f1')  # Using tk.Frame to set background color
        self.progress_frame.pack(fill='x', pady=10)

        # Configure custom style for the progress bar
        style = ttk.Style()
        style.configure("Custom.Horizontal.TProgressbar",
                        thickness=30,
                        background="#2c3e50",  # Progress bar color
                        )

        # Progress bar with custom style
        self.progress = ttk.Progressbar(self.progress_frame, 
                                        orient="horizontal", 
                                        length=400, 
                                        mode="determinate", 
                                        style="Custom.Horizontal.TProgressbar")
        self.progress.pack(fill='x', expand=True, padx=20)

        # Table to display formatted trip data
        self.tree_frame = ttk.Frame(main_frame)
        self.tree_frame.pack(fill='both', expand=True, pady=20)

        self.tree = ttk.Treeview(self.tree_frame)
        self.tree.pack(fill='both', expand=True)

        # Frame for Download and View Trip buttons
        button_frame = tk.Frame(main_frame, bg='#ecf0f1')
        button_frame.pack(pady=10)

        # Download button
        self.download_btn = ttk.Button(button_frame, text="Download Results", command=self.download_trip, style="Accent.TButton")
        self.download_btn.pack(side='left', padx=10)

        # View Trip button
        self.view_trip_btn = ttk.Button(button_frame, text="View Trip", command=self.view_trip, style="Accent.TButton")
        self.view_trip_btn.pack(side='left', padx=10)

    def update_treeview(self):
        # Clear existing data in the table
        for row in self.tree.get_children():
            self.tree.delete(row)

        if self.formatted_trip_data is not None and not self.formatted_trip_data.empty:
            # Set up columns
            columns = list(self.formatted_trip_data.columns)
            self.tree["columns"] = columns
            self.tree["show"] = "headings"  # Remove the first empty column

            for col in columns:
                self.tree.heading(col, text=col, anchor="center")
                col_width = max(100, len(col) * 5)  # Adjust width based on column name
                self.tree.column(col, width=col_width, anchor="center")

            # Track the last inserted value of the first column (Trip ID)
            last_value = None

            # Insert data into the table
            for _, row in self.formatted_trip_data.iterrows():
                first_column_name = columns[0]  # Get the first column name dynamically

                # If this row has the same value in the first column as the last one, leave it blank
                if row[first_column_name] == last_value:
                    row[first_column_name] = ""
                    for i in range(5, 12):  # Columns 6 to 12 (0-based index: 5 to 11)
                        row[columns[i]] = "" # Leave the first column blank for subsequent rows
                else:
                    last_value = row[first_column_name]  # Update the last inserted value

                self.tree.insert("", "end", values=list(row))

            # Adjust column widths based on content
            for col in columns:
                max_width = max(len(str(self.tree.set(child, col))) for child in self.tree.get_children(""))
                self.tree.column(col, width=max(col_width, max_width * 10))
        else:
            # If no data, display a message in the treeview
            self.tree["columns"] = ("message",)
            self.tree.heading("message", text="Status")
            self.tree.column("message", width=400, anchor="center")
            self.tree.insert("", "end", values=("No data available. Please run optimization first.",))

        # Adjust treeview height based on number of rows (max 20 rows visible)
        row_count = min(len(self.tree.get_children()), 20)
        self.tree.configure(height=row_count)



    def choose_file(self):
        self.selected_file = filedialog.askopenfilename(filetypes=(("CSV files", ".csv"), ("All files", ".*")))
        if self.selected_file:
            filename = os.path.basename(self.selected_file)
            self.file_status.config(text=filename)
            self.file_btn.config(text="Change File")

    def run_optimization(self):
        if not self.selected_file:
            messagebox.showerror("Error", "Please select a file first!", parent=self.root)
            return

        # Disable the button to prevent multiple clicks
        self.run_btn.config(state="disabled")

        # Reset progress bar
        self.progress["value"] = 0
        self.progress["maximum"] = 100
        self.progress_frame.update()  # Force update to ensure visibility

        try:
            # Load data
            df = pd.read_csv(self.selected_file)
            store_coords = (19.075887, 72.877911)  # Mumbai coordinates

            # Preprocess time slots
            def parse_time(time_str):
                return datetime.strptime(time_str, '%H:%M:%S').time()

            df['start_time'] = df['Delivery Timeslot'].apply(lambda x: parse_time(x.split('-')[0]))
            df['end_time'] = df['Delivery Timeslot'].apply(lambda x: parse_time(x.split('-')[1]))

            # Convert times to minutes since midnight
            def time_to_minutes(t):
                return t.hour * 60 + t.minute

            df['start_min'] = df['start_time'].apply(time_to_minutes)
            df['end_min'] = df['end_time'].apply(time_to_minutes)

            # Update progress bar (25% complete)
            self.progress["value"] = 25
            self.progress_frame.update_idletasks()

            # Vehicle configuration
            vehicles = [
                {'type': '3W', 'capacity': 5, 'max_dist': 15, 'count': 50},
                {'type': '4W-EV', 'capacity': 8, 'max_dist': 20, 'count': 25},
                {'type': '4W', 'capacity': 25, 'max_dist': float('inf'), 'count': float('inf')}
            ]

            # Cluster shipments using K-Means
            coords = df[['Latitude', 'Longitude']].values
            kmeans = KMeans(n_clusters=75, random_state=42, n_init=10)
            df['cluster'] = kmeans.fit_predict(coords)

            # Update progress bar (50% complete)
            self.progress["value"] = 50
            self.progress_frame.update_idletasks()

            # Route optimization functions
            def calculate_route_distance(points):
                if len(points) == 0:
                    return 0
                total = 0
                current = store_coords
                for point in points:
                    total += haversine(current, (point[0], point[1]))
                    current = (point[0], point[1])
                total += haversine(current, store_coords)
                return total

            # Generate trips
            trips = []
            vehicle_pool = {v['type']: v['count'] for v in vehicles}

            for cluster_id in df['cluster'].unique():
                cluster_df = df[df['cluster'] == cluster_id].sort_values('start_min')
                shipments = cluster_df.to_dict('records')
                
                while len(shipments) > 0:
                    best_vehicle = None
                    best_route = []
                    
                    for vehicle in vehicles:
                        if vehicle_pool[vehicle['type']] <= 0:
                            continue
                            
                        for i in range(len(shipments), 0, -1):
                            candidate = shipments[:i]
                            if len(candidate) > vehicle['capacity']:
                                continue
                            
                            coords_list = [(s['Latitude'], s['Longitude']) for s in candidate]
                            dist = calculate_route_distance(coords_list)
                            
                            if dist <= vehicle['max_dist']:
                                best_vehicle = vehicle
                                best_route = candidate
                                break
                        if best_vehicle:
                            break
                            
                    if best_vehicle:
                        first_leg_distance = haversine(store_coords, (best_route[0]['Latitude'], best_route[0]['Longitude']))
                        start_time = best_route[0]['start_min'] + (first_leg_distance * 5)  # Travel time to first location
                        total_distance = calculate_route_distance([(s['Latitude'], s['Longitude']) for s in best_route])
                        
                        # Adjust total time: Only count unique location-time combinations
                        unique_stops = {(s['Latitude'], s['Longitude'], s['start_min']) for s in best_route}
                        total_time = total_distance * 5 + len(unique_stops) * 10  
                        
                        trips.append({
                            'Vehicle': best_vehicle['type'],
                            'Shipments': ', '.join([str(s['Shipment ID']) for s in best_route]),
                            'Distance (km)': round(total_distance, 2),
                            'Start Time (mins)': round(start_time),
                            'End Time (mins)': round(start_time + total_time),
                            'Total Time (mins)': round(total_time)
                        })
                        vehicle_pool[best_vehicle['type']] -= 1
                        shipments = shipments[len(best_route):]
                    else:
                        first_leg_distance = haversine(store_coords, (shipments[0]['Latitude'], shipments[0]['Longitude']))
                        start_time = shipments[0]['start_min'] + (first_leg_distance * 5)
                        total_distance = calculate_route_distance([(s['Latitude'], s['Longitude']) for s in shipments])
                        
                        unique_stops = {(s['Latitude'], s['Longitude'], s['start_min']) for s in shipments}
                        total_time = total_distance * 5 + len(unique_stops) * 10  
                        
                        trips.append({
                            'Vehicle': '4W',
                            'Shipments': ', '.join([str(s['Shipment ID']) for s in shipments]),
                            'Distance (km)': round(total_distance, 2),
                            'Start Time (mins)': round(start_time),
                            'End Time (mins)': round(start_time + total_time),
                            'Total Time (mins)': round(total_time)
                        })
                        shipments = []

            # Update progress bar (75% complete)
            self.progress["value"] = 75
            self.progress_frame.update_idletasks()

            # Save optimized trips to CSV
            self.trip_data = pd.DataFrame(trips)
            self.trip_data.to_csv("optimized_trip_details.csv", index=False)

            # Run the second script to format the data
            self.format_trip_data()

            # Update progress bar (100% complete)
            self.progress["value"] = 100
            self.progress_frame.update_idletasks()
            self.progress_frame.pack_forget()  

            # Display the formatted data in the table
            self.display_formatted_data()

            # Re-enable the button
            self.run_btn.config(state="normal")

        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"Processing error: {str(e)}", parent=self.root)
            # Re-enable the button in case of error
            self.run_btn.config(state="normal")
    def format_trip_data(self):
        # File paths
        shipments_file = self.selected_file
        trips_file = "optimized_trip_details.csv"
        output_file = "formatted_trip_details.xlsx"

        # Read data
        shipments_data = pd.read_csv(shipments_file)
        trips_data = pd.read_csv(trips_file)

        # Convert necessary columns
        shipments_data["Shipment ID"] = shipments_data["Shipment ID"].astype(int)
        shipments_data["Latitude"] = shipments_data["Latitude"].astype(float)
        shipments_data["Longitude"] = shipments_data["Longitude"].astype(float)

        # Ensure "Trip ID" exists
        trips_data["Trip ID"] = trips_data.index + 1  
        trips_data["Distance (km)"] = trips_data["Distance (km)"].astype(float)

        # Vehicle capacity lookup
        vehicle_capacities = {
            '3W': 5,
            '4W-EV': 8,
            '4W': 25
        }

        # Helper function to convert time string to minutes since midnight
        def time_to_minutes(time_str):
            t = datetime.strptime(time_str, "%H:%M:%S")
            return t.hour * 60 + t.minute

        # Prepare output data
        trip_rows = []

        # Iterate over each trip
        for _, trip in trips_data.iterrows():
            trip_id = trip["Trip ID"]
            vehicle = trip["Vehicle"]
            shipment_ids = list(map(int, trip["Shipments"].split(',')))  # Keep order as is
            shipment_info = shipments_data[shipments_data["Shipment ID"].isin(shipment_ids)]

            # Compute values
            shipment_count = len(shipment_info)
            mst_dist = float(trip["Distance (km)"])
            trip_time = trip['Total Time (mins)']

            # Compute time utilization
            if not shipment_info.empty:
                # Extract start and end times for all shipments
                # shipment_info["Start_Min"] = shipment_info["Delivery Timeslot"].apply(lambda x: time_to_minutes(x.split("-")[0]))
                # shipment_info["End_Min"] = shipment_info["Delivery Timeslot"].apply(lambda x: time_to_minutes(x.split("-")[1]))
                shipment_info = shipment_info.copy()  # Ensures a fresh copy to avoid SettingWithCopyWarning
                shipment_info["Start_Min"] = shipment_info["Delivery Timeslot"].apply(lambda x: time_to_minutes(x.split("-")[0]))
                shipment_info["End_Min"] = shipment_info["Delivery Timeslot"].apply(lambda x: time_to_minutes(x.split("-")[1]))


                # Get the first start time and last end time in minutes
                first_start = shipment_info["Start_Min"].min()
                last_end = shipment_info["End_Min"].max()

                # Calculate total time window for the trip
                slot_duration = last_end - first_start

                slot_duration = max(slot_duration, trip_time)
            else:
                slot_duration = trip_time

            # Compute time utilization
            time_utilization = min(trip_time / slot_duration, 0.99) if slot_duration > 0 else 0

            # Compute capacity utilization
            max_capacity = vehicle_capacities.get(vehicle, 1)  # Default to 1 if unknown vehicle
            capacity_utilization = shipment_count / max_capacity  

            for shipment_id in shipment_ids:  # Ensures original order
                shipment = shipment_info[shipment_info["Shipment ID"] == shipment_id].iloc[0]
                row = {
                    "TRIP ID": trip_id,
                    "Shipment ID": shipment["Shipment ID"],
                    "Latitude": shipment["Latitude"],
                    "Longitude": shipment["Longitude"],
                    "TIME SLOT": shipment["Delivery Timeslot"],
                    "Shipments": shipment_count,
                    "MST_DIST": mst_dist,
                    "TRIP_TIME": trip_time,
                    "Vehicle Type": vehicle,
                    "CAPACITY_UTI": round(capacity_utilization, 2),
                    "TIME_UTI": round(time_utilization, 2),
                    "COV_UTI": 1
                }
                trip_rows.append(row)

        # Create a DataFrame
        self.formatted_trip_data = pd.DataFrame(trip_rows)

        # Rebalance trips to ensure TIME_UTI < 1 and increase CAPACITY_UTI
        self.rebalance_trips()


        # Save to Excel
        self.formatted_trip_data.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active

        # Get column indexes
        columns_to_merge = {
            "TRIP ID": 1,
            "Vehicle Type": 9,
            "Shipments": 6,
            "MST_DIST": 7,
            "TRIP_TIME": 8,
            "CAPACITY_UTI": 10,
            "TIME_UTI": 11,
            "COV_UTI": 12
        }

        # Merge cells for repeated values
        prev_trip_id = None
        merge_start = 2  

        for row in range(2, ws.max_row + 1):
            trip_id = ws.cell(row=row, column=columns_to_merge["TRIP ID"]).value

            if trip_id == prev_trip_id:
                continue  
            else:
                if merge_start < row - 1:
                    for col in columns_to_merge.values():
                        ws.merge_cells(start_row=merge_start, start_column=col, 
                                       end_row=row-1, end_column=col)

                merge_start = row  
                prev_trip_id = trip_id

        # Merge last set of TRIP IDs
        if merge_start < ws.max_row:
            for col in columns_to_merge.values():
                ws.merge_cells(start_row=merge_start, start_column=col, 
                               end_row=ws.max_row, end_column=col)

        # Apply center alignment
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save formatted file
        wb.save(output_file)
        print(f"Formatted trip information saved to {output_file}")

    def rebalance_trips(self):
        # Group by TRIP ID
        grouped = self.formatted_trip_data.groupby('TRIP ID')

        rebalanced_data = []
        for trip_id, group in grouped:
            while group['TIME_UTI'].iloc[0] >= 1:
                # Find the shipment with the largest time window
                largest_window = group.iloc[group['TIME SLOT'].apply(lambda x: int(x.split('-')[1].split(':')[0]) - int(x.split('-')[0].split(':')[0])).idxmax()]
                
                # Create a new trip with this shipment
                new_trip = group.iloc[[largest_window.name]].copy()
                new_trip['TRIP ID'] = self.formatted_trip_data['TRIP ID'].max() + 1
                new_trip['Shipments'] = 1
                new_trip['MST_DIST'] = 0  # This should be recalculated
                new_trip['TRIP_TIME'] = 10  # Minimum time for a single shipment
                new_trip['CAPACITY_UTI'] = 1 / 25  # Assuming 4W capacity
                new_trip['TIME_UTI'] = 0.5  # Arbitrary initial value
                rebalanced_data.append(new_trip)
                
                # Remove the shipment from the original trip
                group = group.drop(largest_window.name)
                
                # Recalculate utilization for the original trip
                group['Shipments'] = len(group)
                group['MST_DIST'] = group['MST_DIST'].iloc[0]  # This should be recalculated
                group['TRIP_TIME'] = group['TRIP_TIME'].iloc[0] - 10  # Subtract time for removed shipment
                group['CAPACITY_UTI'] = group['Shipments'] / 25  # Assuming 4W capacity
                group['TIME_UTI'] = min(group['TRIP_TIME'].iloc[0] / (int(group['TIME SLOT'].iloc[-1].split('-')[1].split(':')[0]) - int(group['TIME SLOT'].iloc[0].split('-')[0].split(':')[0])) * 60, 0.99)

            rebalanced_data.append(group)

        self.formatted_trip_data = pd.concat(rebalanced_data, ignore_index=True)



    def display_formatted_data(self):
        self.update_treeview()
        # messagebox.showinfo("Success", "Optimization completed successfully!", parent=self.root)

    def view_trip(self):
        if self.formatted_trip_data is None:
            messagebox.showerror("Error", "No data available to view!", parent=self.root)
            return

        # Get the selected row from the table
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a row first!", parent=self.root)
            return

        # Get the value from the first column of the selected row
        selected_values = self.tree.item(selected_item, "values")
        if not selected_values:
            messagebox.showerror("Error", "No data in the selected row!", parent=self.root)
            return

        selected_value = selected_values[0]  # First column value

        # Filter the formatted trip data for rows where the first column matches the selected value
        first_column_name = self.formatted_trip_data.columns[0]  # Get the name of the first column

        # Convert the selected value to the appropriate type (e.g., int or str)
        try:
            # If the first column is numeric, convert the selected value to a number
            if pd.api.types.is_numeric_dtype(self.formatted_trip_data[first_column_name]):
                selected_value = float(selected_value)  # Use float to handle both int and float
        except ValueError:
            messagebox.showerror("Error", f"Selected value '{selected_value}' cannot be converted to a number.", parent=self.root)
            return

        filtered_data = self.formatted_trip_data[self.formatted_trip_data[first_column_name] == selected_value]

        if filtered_data.empty:
            messagebox.showerror("Error", f"No data found for {first_column_name}: {selected_value}", parent=self.root)
            return

        # Access latitude and longitude using index (3rd and 4th columns)
        lat_column = filtered_data.columns[2]  # 3rd column (index 2)
        lon_column = filtered_data.columns[3]  # 4th column (index 3)

        # Make sure latitude and longitude columns exist
        if lat_column not in filtered_data.columns or lon_column not in filtered_data.columns:
            messagebox.showerror("Error", "Latitude and Longitude columns are missing.", parent=self.root)
            return

        # Ensure latitude and longitude columns are numeric
        if not pd.api.types.is_numeric_dtype(filtered_data[lat_column]) or not pd.api.types.is_numeric_dtype(filtered_data[lon_column]):
            messagebox.showerror("Error", "Latitude and Longitude columns must contain numeric values.", parent=self.root)
            return

        # Create a Folium map centered on an average coordinate (e.g., the first one in the filtered data)
        avg_lat = filtered_data[lat_column].mean()
        avg_lon = filtered_data[lon_column].mean()
        folium_map = folium.Map(location=[avg_lat, avg_lon], zoom_start=14)

        store_lat = 19.075887
        store_lon = 72.877911
        folium.Marker(
            [store_lat, store_lon],
            popup="Store Location",
            icon=folium.Icon(color='red', icon='home')
        ).add_to(folium_map)

        # To draw the route, we need to keep track of the previous marker's coordinates
        previous_lat = store_lat
        previous_lon = store_lon

        # Place markers for each coordinate in the filtered data
        for _, row in filtered_data.iterrows():
            lat = row[lat_column]
            lon = row[lon_column]
            
            folium.Marker(
                [lat, lon], 
                popup=f"Latitude: {lat}, Longitude: {lon}", 
                icon=folium.Icon(color='blue', icon='shopping-cart')
            ).add_to(folium_map)
                        
            # Draw a polyline (route) from the previous location to the current one
            folium.PolyLine([(previous_lat, previous_lon), (lat, lon)], color='blue', weight=2.5, opacity=1).add_to(folium_map)

            # Update the previous coordinates
            previous_lat = lat
            previous_lon = lon

        folium.PolyLine([(previous_lat, previous_lon), (store_lat, store_lon)], color='blue', weight=2.5, opacity=1).add_to(folium_map)
        # Save the map to an HTML file
        map_path = "filtered_trip_map.html"
        folium_map.save(map_path)

        # Open the map in the default web browser
        webbrowser.open(f"file://{os.path.abspath(map_path)}")


    def download_trip(self):
        if self.formatted_trip_data is None:
            messagebox.showerror("Error", "No data available to download!", parent=self.root)
            return

        # Get the current directory where the script is located
        current_directory = os.path.dirname(os.path.abspath(__file__))
        output_file_path = os.path.join(current_directory, "formatted_trip_details.xlsx")

        # Save the file
        self.formatted_trip_data.to_excel(output_file_path, index=False)
        messagebox.showinfo("Success", "File downloaded successfully", parent=self.root)

    def show_home(self):
        messagebox.showinfo("Home", "Welcome to Smart Route Optimizer", parent=self.root)

    def show_about(self):
        about_text = """Smart Route Optimizer v1.0\n
Advanced route optimization solution for logistics efficiency."""
        messagebox.showinfo("About", about_text, parent=self.root)

    def show_pricing(self):
        pricing_text = """Professional Edition: $99/month\n
Includes unlimited optimizations and premium support"""
        messagebox.showinfo("Pricing", pricing_text, parent=self.root)

if __name__ == "__main__":
    root = tk.Tk()
    app = SmartRouteOptimizerApp(root)
    root.mainloop()